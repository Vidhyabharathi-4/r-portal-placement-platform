import io
import os
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
from fastapi import Depends, FastAPI, File, HTTPException, Query, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session, joinedload
from .ats import calculate_ats_match, extract_text_from_file, parse_jd_text
from .audit import log_action
from .config import get_settings
from .database import Base, SessionLocal, ensure_sqlite_schema, engine, get_db
from .dependencies import get_current_user, require_roles
from .models import Application, ApplicationStatus, AuditLog, Company, DriveStatus, Notification, PlacementDrive, PlacementStatus, PlacementTeamDriveAssignment, PlacementTeamMember, RecruiterContact, RecruiterStatus, Role, Student, User
from .schemas import (
    ApplicationCreate, ApplicationOut, ApplicationStatusUpdate, AuditOut,
    ATSBulkMatchOut, ATSMatchItem, CompanyBase, CompanyCardOut, CompanyCreate,
    CompanyDetailsOut, CompanyOut, CompanyStatusUpdate, CompanyUpdate,
    DashboardOut, DriveCreate, DriveOut, DriveUpdate, LoginRequest,
    NotificationOut, PasswordChange, ProfileUpdate, RecruiterContactBase,
    RecruiterContactCreate, RecruiterContactOut, RecruiterContactUpdate,
    RecruiterMetricsOut, RecruitersOverviewOut, ReportsOut, SettingsUpdate,
    StudentCreate, StudentOut, TeamMemberCreate, TeamMemberOut,
    TeamMemberUpdate, Token, UserCreate, UserOut, UserStatusUpdate,
)
from .security import create_access_token, hash_password, verify_password

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

Base.metadata.create_all(bind=engine)
ensure_sqlite_schema()


def seed_default_users():
    default_accounts = [
        {
            "full_name": "sivasubramaniyam",
            "email": "sivasubramaniyam@gmail.com",
            "password": "SS@Rathinam",
            "role": Role.ADMIN,
        },
        {
            "full_name": "Jeyakkanan",
            "email": "jeyakkanan@gmail.com",
            "password": "Jk@Rathinam",
            "role": Role.MANAGER,
        },
        {
            "full_name": "Swetha",
            "email": "swetha@gmail.com",
            "password": "Swetha@Rathinam",
            "role": Role.LEAD,
        },
        {
            "full_name": "System Admin",
            "email": "admin@rportal.com",
            "password": "admin123",
            "role": Role.ADMIN,
        },
    ]
    with SessionLocal() as db:
        for acc in default_accounts:
            email_clean = acc["email"].strip().lower()
            existing = db.scalar(select(User).where(User.email == email_clean))
            if not existing:
                u = User(
                    full_name=acc["full_name"],
                    email=email_clean,
                    password_hash=hash_password(acc["password"]),
                    role=acc["role"],
                    is_active=True,
                    preferences={},
                )
                db.add(u)
            else:
                existing.password_hash = hash_password(acc["password"])
                existing.role = acc["role"]
                existing.is_active = True
        db.commit()


def seed_default_recruiter_contacts():
    with SessionLocal() as db:
        companies = db.scalars(select(Company)).all()
        for comp in companies:
            existing_count = db.scalar(select(func.count()).select_from(RecruiterContact).where(RecruiterContact.company_id == comp.id)) or 0
            if existing_count == 0 and comp.contact_name:
                recruiter = RecruiterContact(
                    company_id=comp.id,
                    name=comp.contact_name,
                    designation=comp.contact_designation or "Head of Campus Recruitment",
                    email=comp.contact_email or f"campus@{comp.name.lower().replace(' ', '')}.com",
                    phone=comp.contact_phone or "+91 98765 43210",
                    department="University Relations / Talent Acquisition",
                    status="ACTIVE",
                    last_contacted=comp.last_contacted_at or comp.updated_at,
                    notes=f"Primary campus recruitment coordinator for {comp.name}."
                )
                db.add(recruiter)
        db.commit()


seed_default_users()
seed_default_recruiter_contacts()
app = FastAPI(title="R-PORTAL API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=get_settings().origins, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])
UPLOAD_DIR = Path("/tmp/uploads" if os.getenv("VERCEL") else "uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


DEFAULT_USER_PREFERENCES = {
    "theme": "system",
    "table_density": "comfortable",
    "default_page": "/dashboard",
    "default_export_format": "CSV",
    "default_print_orientation": "portrait",
    "notifications": {
        "application_updates": True,
        "drive_updates": True,
        "recruiter_updates": True,
        "student_updates": True,
        "system_updates": True,
    },
}


def audit_commit(db: Session):
    db.commit()


def should_receive_notification(recipient: User, notification_type: str) -> bool:
    prefs = recipient.preferences or {}
    notif_prefs = prefs.get("notifications", {})
    if not isinstance(notif_prefs, dict):
        notif_prefs = {}

    mapping = {
        "APPLICATION_CREATED": "application_updates",
        "APPLICATION_STATUS_CHANGED": "application_updates",
        "DRIVE_CREATED": "drive_updates",
        "DRIVE_STATUS_CHANGED": "drive_updates",
        "DRIVE_DELETED": "drive_updates",
        "DRIVE_ASSIGNED": "drive_updates",
        "COMPANY_CREATED": "recruiter_updates",
        "COMPANY_STATUS_CHANGED": "recruiter_updates",
        "RECRUITER_CREATED": "recruiter_updates",
        "RECRUITER_STATUS_CHANGED": "recruiter_updates",
        "RECRUITER_DELETED": "recruiter_updates",
        "STUDENT_CREATED": "student_updates",
        "STUDENT_UPDATED": "student_updates",
        "STUDENT_DELETED": "student_updates",
        "STUDENTS_IMPORTED": "student_updates",
        "USER_REGISTERED": "system_updates",
        "TEAM_IMPORTED": "system_updates",
        "TEAM_MEMBER_ADDED": "system_updates",
        "TEAM_MEMBER_UPDATED": "system_updates",
    }
    category = mapping.get(notification_type, "system_updates")
    return notif_prefs.get(category, True)


def create_notification(db: Session, recipient_id: int, title: str, message: str, notification_type: str, entity_type: str | None = None, entity_id: int | str | None = None):
    recipient = db.get(User, recipient_id)
    if recipient and not should_receive_notification(recipient, notification_type):
        return
    db.add(Notification(recipient_id=recipient_id, title=title, message=message, notification_type=notification_type, entity_type=entity_type, entity_id=str(entity_id) if entity_id is not None else None))


def notify_admins_and_managers(db: Session, title: str, message: str, notification_type: str, entity_type: str | None = None, entity_id: int | str | None = None, exclude_user_id: int | None = None):
    recipients = db.scalars(select(User).where(User.is_active.is_(True), User.role.in_([Role.ADMIN, Role.MANAGER]))).all()
    for recipient in recipients:
        if exclude_user_id is not None and recipient.id == exclude_user_id:
            continue
        create_notification(db, recipient.id, title, message, notification_type, entity_type, entity_id)


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.post("/api/auth/register", response_model=Token, status_code=status.HTTP_201_CREATED)
def register(payload: UserCreate, db: Session = Depends(get_db)):
    if db.scalar(select(User).where(User.email == payload.email)):
        raise HTTPException(409, "An account with this email already exists")
    user = User(full_name=payload.full_name, email=str(payload.email).lower(), password_hash=hash_password(payload.password), role=payload.role)
    db.add(user); db.flush()
    log_action(db, user, "CREATE", "user", user.id, {"role": user.role.value, "email": user.email})
    notify_admins_and_managers(db, "New User Registered", f"{user.full_name} ({user.role.value}) created an account.", "USER_REGISTERED", "user", user.id, exclude_user_id=user.id)
    audit_commit(db); db.refresh(user)
    return Token(access_token=create_access_token(str(user.id)), user=user)


@app.post("/api/auth/login", response_model=Token)
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.scalar(select(User).where(User.email == str(payload.email).lower()))
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.is_active:
        raise HTTPException(403, "Account is inactive")
    log_action(db, user, "LOGIN", "user", user.id)
    audit_commit(db)
    return Token(access_token=create_access_token(str(user.id)), user=user)


@app.get("/api/auth/me", response_model=UserOut)
def me(current: User = Depends(get_current_user)):
    return current


@app.get("/api/settings")
def get_user_settings(current: User = Depends(get_current_user)):
    user_prefs = current.preferences or {}
    merged_prefs = {**DEFAULT_USER_PREFERENCES, **user_prefs}
    if "notifications" in user_prefs and isinstance(user_prefs["notifications"], dict):
        merged_prefs["notifications"] = {
            **DEFAULT_USER_PREFERENCES["notifications"],
            **user_prefs["notifications"],
        }
    return {
        "user": current,
        "preferences": merged_prefs,
    }


@app.patch("/api/profile", response_model=UserOut)
def update_profile(payload: ProfileUpdate, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    cleaned_name = payload.full_name.strip()
    if len(cleaned_name) < 2:
        raise HTTPException(422, "Full name must be at least 2 characters long.")
    user = db.get(User, current.id)
    if not user:
        user = db.merge(current)
    old_name = user.full_name
    user.full_name = cleaned_name
    log_action(db, user, "UPDATE_PROFILE", "user", user.id, {"from": old_name, "to": cleaned_name})
    audit_commit(db)
    db.refresh(user)
    return user


@app.patch("/api/settings", response_model=UserOut)
def update_settings(payload: SettingsUpdate, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    user = db.get(User, current.id)
    if not user:
        user = db.merge(current)
    existing = dict(user.preferences or DEFAULT_USER_PREFERENCES)
    updates = payload.model_dump(exclude_unset=True)
    
    if "notifications" in updates and updates["notifications"] is not None:
        existing_notifs = dict(existing.get("notifications") or DEFAULT_USER_PREFERENCES["notifications"])
        existing_notifs.update(updates["notifications"])
        existing["notifications"] = existing_notifs
        del updates["notifications"]
        
    for k, v in updates.items():
        if v is not None:
            existing[k] = v

    user.preferences = existing
    log_action(db, user, "UPDATE_SETTINGS", "user", user.id, {"updated_keys": list(payload.model_dump(exclude_unset=True).keys())})
    audit_commit(db)
    db.refresh(user)
    return user


@app.post("/api/auth/change-password")
def change_password(payload: PasswordChange, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    user = db.get(User, current.id)
    if not user:
        user = db.merge(current)
    if not verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=422, detail="New password must be at least 8 characters.")
    user.password_hash = hash_password(payload.new_password)
    log_action(db, user, "CHANGE_PASSWORD", "user", user.id)
    audit_commit(db)
    return {"message": "Password changed successfully."}


@app.patch("/api/users/{user_id}/status", response_model=UserOut)
def update_user_status(user_id: int, payload: UserStatusUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if target.id == current.id and not payload.is_active:
        raise HTTPException(400, "You cannot deactivate your own account.")
    target.is_active = payload.is_active
    log_action(db, current, "UPDATE_USER_STATUS", "user", target.id, {"email": target.email, "is_active": payload.is_active})
    audit_commit(db)
    db.refresh(target)
    return target


@app.get("/api/users", response_model=list[UserOut])
def list_users(include_inactive: bool = False, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    stmt = select(User)
    if not include_inactive:
        stmt = stmt.where(User.is_active.is_(True))
    return db.scalars(stmt.order_by(User.full_name)).all()


@app.get("/api/dashboard", response_model=DashboardOut)
def dashboard(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    drives = db.scalars(select(PlacementDrive).options(joinedload(PlacementDrive.company)).order_by(PlacementDrive.created_at.desc()).limit(5)).unique().all()
    activity = db.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(8)).all()
    total_students=db.scalar(select(func.count()).select_from(Student)) or 0; eligible=db.scalar(select(func.count()).select_from(Student).where(Student.is_eligible.is_(True))) or 0; placed=db.scalar(select(func.count()).select_from(Student).where(Student.placement_status==PlacementStatus.PLACED)) or 0
    return DashboardOut(total_students=total_students, eligible_students=eligible, placed_students=placed, placement_percentage=round((placed/total_students)*100,2) if total_students else 0, active_drives=db.scalar(select(func.count()).select_from(PlacementDrive).where(PlacementDrive.status == DriveStatus.OPEN,PlacementDrive.is_archived.is_(False))) or 0, total_companies=db.scalar(select(func.count()).select_from(Company)) or 0, total_applications=db.scalar(select(func.count()).select_from(Application)) or 0, offers=db.scalar(select(func.count()).select_from(Application).where(Application.status == ApplicationStatus.OFFERED)) or 0, recent_drives=drives, recent_activity=activity)


@app.get("/api/companies", response_model=list[CompanyCardOut])
def list_companies(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    companies = db.scalars(select(Company).order_by(Company.name)).all()
    results = []
    for c in companies:
        recruiters_count = db.scalar(select(func.count()).select_from(RecruiterContact).where(RecruiterContact.company_id == c.id)) or 0
        drives = db.scalars(select(PlacementDrive).where(PlacementDrive.company_id == c.id).order_by(PlacementDrive.created_at.desc())).all()
        drive_ids = [d.id for d in drives]
        
        placed_students_count = db.scalar(
            select(func.count()).select_from(Student).where(Student.placed_company_id == c.id)
        ) or 0
        
        latest_drive = drives[0] if drives else None
        
        results.append(CompanyCardOut(
            id=c.id,
            name=c.name,
            website=c.website,
            industry=c.industry,
            location=c.location,
            address=c.address,
            description=c.description,
            contact_name=c.contact_name,
            contact_email=c.contact_email,
            contact_phone=c.contact_phone,
            contact_designation=c.contact_designation,
            logo_url=c.logo_url,
            notes=c.notes,
            last_contacted_at=c.last_contacted_at,
            recruiter_status=c.recruiter_status,
            created_at=c.created_at,
            updated_at=c.updated_at,
            recruiter_count=recruiters_count,
            primary_contact=c.contact_name,
            primary_email=c.contact_email,
            primary_phone=c.contact_phone,
            primary_designation=c.contact_designation or "HR Manager",
            total_drives=len(drives),
            latest_drive_title=latest_drive.title if latest_drive else None,
            latest_drive_date=latest_drive.drive_date if latest_drive else None,
            latest_drive_status=latest_drive.status.value if latest_drive else None,
            students_placed_count=placed_students_count,
        ))
    return results


@app.post("/api/companies", response_model=CompanyOut, status_code=201)
def create_company(payload: CompanyCreate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    if db.scalar(select(Company).where(Company.name.ilike(payload.name.strip()))):
        raise HTTPException(409, f"A company with the name '{payload.name}' already exists.")
    
    company_data = payload.model_dump()
    company = Company(**company_data)
    if not company.last_contacted_at:
        company.last_contacted_at = datetime.now(timezone.utc)
    db.add(company)
    db.flush()
    
    # If primary recruiter contact details are provided, auto-create a recruiter contact
    if company.contact_email:
        recruiter = RecruiterContact(
            company_id=company.id,
            name=company.contact_name or f"{company.name} Campus HR",
            designation=company.contact_designation or "HR Manager",
            email=company.contact_email,
            phone=company.contact_phone,
            status="ACTIVE",
            last_contacted=datetime.now(timezone.utc),
            notes=f"Primary contact for {company.name}"
        )
        db.add(recruiter)
        db.flush()
    
    log_action(db, current, "CREATE", "company", company.id, {"name": company.name, "status": company.recruiter_status.value})
    notify_admins_and_managers(db, "New Company Added", f"{company.name} was added to the platform.", "COMPANY_CREATED", "company", company.id, exclude_user_id=current.id)
    audit_commit(db)
    db.refresh(company)
    return company


@app.put("/api/companies/{company_id}", response_model=CompanyOut)
@app.patch("/api/companies/{company_id}", response_model=CompanyOut)
def update_company(company_id: int, payload: CompanyUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    company = db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")
    
    update_data = payload.model_dump(exclude_unset=True)
    if "name" in update_data and update_data["name"] and update_data["name"].strip() != company.name:
        existing = db.scalar(select(Company).where(Company.name.ilike(update_data["name"].strip()), Company.id != company_id))
        if existing:
            raise HTTPException(409, f"Another company with name '{update_data['name']}' already exists.")
    
    old_status = company.recruiter_status.value
    for key, value in update_data.items():
        if value is not None:
            setattr(company, key, value)
    
    company.last_contacted_at = datetime.now(timezone.utc)
    log_action(db, current, "UPDATE", "company", company.id, {"name": company.name, "status": company.recruiter_status.value})
    if old_status != company.recruiter_status.value:
        notify_admins_and_managers(db, "Company Status Changed", f"{company.name} status changed from {old_status} to {company.recruiter_status.value}.", "COMPANY_STATUS_CHANGED", "company", company.id, exclude_user_id=current.id)
    
    audit_commit(db)
    db.refresh(company)
    return company


@app.delete("/api/companies/{company_id}", status_code=204)
def delete_company(company_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    company = db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")
    log_action(db, current, "DELETE", "company", company.id, {"name": company.name})
    notify_admins_and_managers(db, "Company Deleted", f"Company '{company.name}' was deleted.", "COMPANY_DELETED", "company", company.id, exclude_user_id=current.id)
    db.delete(company)
    audit_commit(db)


@app.get("/api/drives", response_model=list[DriveOut])
def list_drives(status_filter: DriveStatus | None = None, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    query = select(PlacementDrive).options(joinedload(PlacementDrive.company)).order_by(PlacementDrive.created_at.desc())
    if status_filter:
        query = query.where(PlacementDrive.status == status_filter)
    return db.scalars(query).unique().all()


@app.get("/api/drives/{drive_id}", response_model=DriveOut)
def get_drive(drive_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    drive = db.scalar(select(PlacementDrive).options(joinedload(PlacementDrive.company)).where(PlacementDrive.id == drive_id))
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    return drive


@app.post("/api/drives", response_model=DriveOut, status_code=201)
def create_drive(payload: DriveCreate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    company = db.get(Company, payload.company_id)
    if not company:
        raise HTTPException(404, "Company not found")
    
    drive_data = payload.model_dump()
    drive = PlacementDrive(**drive_data, created_by_id=current.id)
    db.add(drive)
    db.flush()
    
    log_action(db, current, "CREATE", "placement_drive", drive.id, {"title": drive.title, "company": company.name, "status": drive.status.value})
    notify_admins_and_managers(db, "New Placement Drive", f"{drive.title} was created for {company.name}.", "DRIVE_CREATED", "placement_drive", drive.id, exclude_user_id=current.id)
    audit_commit(db)
    db.refresh(drive)
    return db.scalar(select(PlacementDrive).options(joinedload(PlacementDrive.company)).where(PlacementDrive.id == drive.id))


@app.put("/api/drives/{drive_id}", response_model=DriveOut)
@app.patch("/api/drives/{drive_id}", response_model=DriveOut)
def update_drive(drive_id: int, payload: DriveUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    drive = db.get(PlacementDrive, drive_id)
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    if not db.get(Company, payload.company_id):
        raise HTTPException(404, "Company not found")
    
    previous_status = drive.status.value
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(drive, key, value)
    
    log_action(db, current, "STATUS_CHANGE" if previous_status != drive.status.value else "UPDATE", "placement_drive", drive.id, {"title": drive.title, "from": previous_status, "to": drive.status.value})
    if previous_status != drive.status.value:
        notify_admins_and_managers(db, "Drive Status Updated", f"{drive.title} status changed from {previous_status} to {drive.status.value}.", "DRIVE_STATUS_CHANGED", "placement_drive", drive.id, exclude_user_id=current.id)
    
    audit_commit(db)
    return db.scalar(select(PlacementDrive).options(joinedload(PlacementDrive.company)).where(PlacementDrive.id == drive.id))


@app.delete("/api/drives/{drive_id}", status_code=204)
def delete_drive(drive_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    drive = db.get(PlacementDrive, drive_id)
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    log_action(db, current, "DELETE", "placement_drive", drive.id, {"title": drive.title})
    notify_admins_and_managers(db, "Drive Deleted", f"Placement drive '{drive.title}' was deleted.", "DRIVE_DELETED", "placement_drive", drive.id, exclude_user_id=current.id)
    db.delete(drive)
    audit_commit(db)


@app.post("/api/drives/{drive_id}/jd-upload")
def upload_drive_jd_file(
    drive_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    drive = db.get(PlacementDrive, drive_id)
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    
    filename = file.filename or "jd_document"
    contents = file.file.read()
    if not contents:
        raise HTTPException(400, "Empty document uploaded.")
    
    # Save file to uploads directory
    safe_name = f"jd-{drive.id}-{uuid4().hex[:8]}-{Path(filename).name}"
    target_path = UPLOAD_DIR / safe_name
    with target_path.open("wb") as f:
        f.write(contents)
    
    # Extract text and parse requirements
    extracted_text = extract_text_from_file(contents, filename)
    parsed = parse_jd_text(extracted_text)
    
    drive.jd_document_path = safe_name
    drive.jd_text = extracted_text
    if parsed.get("required_skills"):
        drive.required_skills = ", ".join(parsed["required_skills"])
    if parsed.get("preferred_skills"):
        drive.preferred_skills = ", ".join(parsed["preferred_skills"])
    if parsed.get("min_cgpa"):
        drive.min_cgpa = parsed["min_cgpa"]
    if parsed.get("max_backlogs") is not None:
        drive.max_backlogs = parsed["max_backlogs"]
    if parsed.get("eligible_departments"):
        drive.departments = ", ".join(parsed["eligible_departments"])
    if parsed.get("package_lpa") and not drive.package_lpa:
        drive.package_lpa = parsed["package_lpa"]
    
    log_action(db, current, "UPLOAD_JD", "placement_drive", drive.id, {"filename": safe_name, "skills": drive.required_skills})
    audit_commit(db)
    
    return {
        "message": "Job description parsed and attached successfully!",
        "filename": safe_name,
        "extracted_text_preview": parsed.get("extracted_text_preview", ""),
        "parsed_requirements": parsed,
        "drive_id": drive.id,
    }


@app.get("/api/drives/{drive_id}/ats-match", response_model=ATSBulkMatchOut)
def get_drive_ats_matches(drive_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    drive = db.scalar(select(PlacementDrive).options(joinedload(PlacementDrive.company)).where(PlacementDrive.id == drive_id))
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    
    students = db.scalars(select(Student).order_by(Student.name)).all()
    existing_applications = {
        app.student_id: app.status.value for app in db.scalars(select(Application).where(Application.drive_id == drive_id)).all() if app.student_id
    }
    
    matches = []
    eligible_count = 0
    high_match_count = 0
    
    for s in students:
        match_data = calculate_ats_match(s, drive)
        has_app = s.id in existing_applications
        app_status = existing_applications.get(s.id)
        
        match_item = ATSMatchItem(
            student_id=s.id,
            student_name=s.name,
            registration_number=s.registration_number,
            department=match_data["department"],
            cgpa=match_data["cgpa"],
            skills=match_data["skills"],
            ats_score=match_data["ats_score"],
            skills_match_pct=match_data["skills_match_pct"],
            academic_match_pct=match_data["academic_match_pct"],
            dept_match_pct=match_data["dept_match_pct"],
            matched_skills=match_data["matched_skills"],
            missing_skills=match_data["missing_skills"],
            matched_preferred_skills=match_data["matched_preferred_skills"],
            is_eligible=match_data["is_eligible"],
            reasons=match_data["reasons"],
            has_applied=has_app,
            application_status=app_status,
        )
        
        if match_item.is_eligible:
            eligible_count += 1
        if match_item.ats_score >= 70:
            high_match_count += 1
            
        matches.append(match_item)
    
    # Sort ranked by ATS Score descending
    matches.sort(key=lambda m: (m.is_eligible, m.ats_score), reverse=True)
    
    req_skills = [s.strip() for s in (drive.required_skills or "").split(",") if s.strip()]
    pref_skills = [s.strip() for s in (drive.preferred_skills or "").split(",") if s.strip()]
    depts = [d.strip() for d in (drive.departments or "").split(",") if d.strip()]
    
    min_cgpa_val = 6.0
    try:
        min_cgpa_val = float(str(drive.min_cgpa).split()[0])
    except Exception:
        min_cgpa_val = 6.0
        
    return ATSBulkMatchOut(
        drive_id=drive.id,
        drive_title=drive.title,
        company_name=drive.company.name if drive.company else "Company",
        required_skills=req_skills,
        preferred_skills=pref_skills,
        min_cgpa=min_cgpa_val,
        max_backlogs=drive.max_backlogs or 0,
        eligible_departments=depts,
        total_candidates=len(students),
        eligible_count=eligible_count,
        high_match_count=high_match_count,
        matches=matches,
    )


@app.post("/api/drives/{drive_id}/ats-shortlist")
def shortlist_candidates_bulk(
    drive_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    drive = db.get(PlacementDrive, drive_id)
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    
    student_ids = payload.get("student_ids", [])
    if not student_ids:
        raise HTTPException(400, "No students selected for shortlisting.")
    
    shortlisted_count = 0
    for sid in student_ids:
        student = db.get(Student, sid)
        if not student:
            continue
        
        app = db.scalar(select(Application).where(Application.drive_id == drive_id, Application.student_id == sid))
        if app:
            app.status = ApplicationStatus.SHORTLISTED
        else:
            app = Application(
                drive_id=drive_id,
                student_id=student.id,
                student_name=student.name,
                student_email=student.email,
                status=ApplicationStatus.SHORTLISTED,
            )
            db.add(app)
        shortlisted_count += 1
    
    log_action(db, current, "ATS_SHORTLIST", "placement_drive", drive.id, {"shortlisted_count": shortlisted_count})
    notify_admins_and_managers(db, "Candidates Shortlisted", f"{shortlisted_count} candidates shortlisted for {drive.title} via ATS matching.", "APPLICATION_STATUS_CHANGED", "placement_drive", drive.id, exclude_user_id=current.id)
    audit_commit(db)
    
    return {
        "message": f"Successfully shortlisted {shortlisted_count} candidate(s) for {drive.title}!",
        "shortlisted_count": shortlisted_count
    }


@app.get("/api/applications", response_model=list[ApplicationOut])
def list_applications(drive_id: int | None = None, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    query = select(Application).options(joinedload(Application.drive).joinedload(PlacementDrive.company)).order_by(Application.created_at.desc())
    if drive_id: query = query.where(Application.drive_id == drive_id)
    return db.scalars(query).unique().all()


@app.post("/api/applications", response_model=ApplicationOut, status_code=201)
def create_application(payload: ApplicationCreate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    if not db.get(PlacementDrive, payload.drive_id): raise HTTPException(404, "Placement drive not found")
    application = Application(**payload.model_dump()); db.add(application); db.flush()
    log_action(db, current, "CREATE", "application", application.id, {"student_email": application.student_email})
    notify_admins_and_managers(db, "New Application Submitted", f"Application submitted for {application.student_name}.", "APPLICATION_CREATED", "application", application.id, exclude_user_id=current.id)
    audit_commit(db)
    return db.scalar(select(Application).options(joinedload(Application.drive).joinedload(PlacementDrive.company)).where(Application.id == application.id))


@app.patch("/api/applications/{application_id}/status", response_model=ApplicationOut)
def update_application_status(application_id: int, payload: ApplicationStatusUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    application = db.get(Application, application_id)
    if not application: raise HTTPException(404, "Application not found")
    before = application.status.value; application.status = payload.status
    log_action(db, current, "STATUS_CHANGE", "application", application.id, {"from": before, "to": payload.status.value})
    notify_admins_and_managers(db, "Application Status Updated", f"Application for {application.student_name} changed to {payload.status.value}.", "APPLICATION_STATUS_CHANGED", "application", application.id, exclude_user_id=current.id)
    if application.student_id:
        student = db.get(Student, application.student_id)
        if student and payload.status == ApplicationStatus.OFFERED:
            student.placement_status = PlacementStatus.PLACED
            if application.drive and application.drive.company_id:
                student.placed_company_id = application.drive.company_id
    audit_commit(db)
    return db.scalar(select(Application).options(joinedload(Application.drive).joinedload(PlacementDrive.company)).where(Application.id == application.id))


@app.post("/api/applications/{application_id}/resume", response_model=ApplicationOut)
def upload_resume(application_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    application = db.get(Application, application_id)
    if not application: raise HTTPException(404, "Application not found")
    if Path(file.filename or "").suffix.lower() not in {".pdf", ".doc", ".docx"}: raise HTTPException(400, "Only PDF, DOC and DOCX files are allowed")
    filename = f"{application.id}-{uuid4().hex}{Path(file.filename).suffix.lower()}"; target = UPLOAD_DIR / filename
    with target.open("wb") as buffer: shutil.copyfileobj(file.file, buffer)
    application.resume_path = filename; log_action(db, current, "UPLOAD", "application_resume", application.id, {"filename": filename}); audit_commit(db)
    return db.scalar(select(Application).options(joinedload(Application.drive).joinedload(PlacementDrive.company)).where(Application.id == application.id))


@app.get("/api/uploads/{filename}")
def download_upload(filename: str, current: User = Depends(get_current_user)):
    safe_name = Path(filename).name; path = UPLOAD_DIR / safe_name
    if not path.exists(): raise HTTPException(404, "File not found")
    return FileResponse(path)


@app.get("/api/audit", response_model=list[AuditOut])
def list_audit(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    return db.scalars(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(200)).all()

def _normalize_excel_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _clean_excel_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def _parse_excel_rows(file_bytes: bytes, filename: str):
    ext = Path(filename).suffix.lower()
    if ext not in {".xlsx", ".xls", ".csv"}:
        raise ValueError("Only .xlsx, .xls, and .csv files are supported.")
        
    if ext == ".csv":
        text_data = None
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                text_data = file_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text_data is None:
            text_data = file_bytes.decode("latin-1", errors="replace")
        import csv
        reader = csv.reader(io.StringIO(text_data))
        rows = [list(row) for row in reader]
    elif ext == ".xlsx":
        if openpyxl is None:
            raise ValueError("Excel import requires openpyxl to be installed.")
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        if not workbook.sheetnames:
            raise ValueError("The Excel file does not contain any sheets.")
        worksheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        if xlrd is None:
            raise ValueError("Excel import requires xlrd to be installed.")
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        rows = [[sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)] for row_index in range(sheet.nrows)]
    
    # Filter completely empty rows
    rows = [row for row in rows if any(_clean_excel_value(c) != "" for c in row)]
    if not rows:
        raise ValueError("The uploaded file contains no data rows.")
    
    # Find the header row by searching known header keywords
    known_headers = {
        "roll no", "roll number", "registration number", "registration no", "reg no", "reg number", 
        "register number", "register no", "regno", "rollno", "student id", "name", "student name", 
        "full name", "candidate name", "department", "dept", "branch", "email", "mail", "phone", 
        "mobile", "contact", "cgpa", "gpa", "placement status", "status", "course", "degree"
    }
    header_row_index = 0
    for idx, row in enumerate(rows[:10]):
        normalized_headers = {_normalize_excel_header(cell) for cell in row if cell is not None}
        if normalized_headers.intersection(known_headers):
            header_row_index = idx
            break
            
    rows = rows[header_row_index:]
    if not rows:
        raise ValueError("No valid header row found in the file.")
    return rows


def _parse_placement_status(raw_value: object) -> PlacementStatus:
    value = _clean_excel_value(raw_value).lower().replace(" ", "")
    mapping = {
        "placed": PlacementStatus.PLACED,
        "selected": PlacementStatus.PLACED,
        "offer": PlacementStatus.PLACED,
        "offered": PlacementStatus.PLACED,
        "hired": PlacementStatus.PLACED,
        "unplaced": PlacementStatus.SEEKING,
        "notplaced": PlacementStatus.SEEKING,
        "available": PlacementStatus.SEEKING,
        "seeking": PlacementStatus.SEEKING,
        "noteligible": PlacementStatus.NOT_ELIGIBLE,
        "ineligible": PlacementStatus.NOT_ELIGIBLE,
        "eligible": PlacementStatus.SEEKING,
    }
    return mapping.get(value, PlacementStatus.SEEKING)


def _coerce_float(value: object) -> float | None:
    if value is None or _clean_excel_value(value) == "":
        return None
    try:
        raw = _clean_excel_value(value).replace(",", "").replace("%", "").strip()
        number = float(raw)
        if number > 10 and number <= 100:  # Percentage converted to 10-scale CGPA
            number = round(number / 10.0, 2)
    except (TypeError, ValueError):
        return None
    return number


@app.get("/api/students/sample-template")
def download_student_template(format: str = Query("xlsx", pattern="^(xlsx|csv)$")):
    headers = ["Registration Number", "Student Name", "Department", "Email", "Phone", "CGPA", "Placement Status", "Skills"]
    sample_rows = [
        ["2026-CSE-101", "Aravind Kumar", "CSE", "aravind.k@rathinam.edu.in", "9876543210", "8.75", "Seeking", "Python, React, SQL"],
        ["2026-IT-102", "Priya Dharshini", "IT", "priya.d@rathinam.edu.in", "9876543211", "9.10", "Placed", "Java, Spring Boot, AWS"],
        ["2026-ECE-103", "Rahul Sharma", "ECE", "rahul.s@rathinam.edu.in", "9876543212", "7.60", "Seeking", "Embedded C, IoT, Python"],
    ]
    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in sample_rows:
            writer.writerow(row)
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="students_sample_template.csv"'}
        )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(headers)
        for row in sample_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="students_sample_template.xlsx"'}
        )


@app.post("/api/students/import")
def import_students(
    file: UploadFile = File(...),
    mode: str = Query("skip", pattern="^(skip|upsert)$"),
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    if not file.filename:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No file selected.")

    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xls", ".csv"}:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx, .xls, and .csv files are supported.")

    try:
        contents = file.file.read()
        rows = _parse_excel_rows(contents, file.filename)
    except Exception as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Invalid import file: {exc}") from exc

    headers = [_clean_excel_value(cell) for cell in rows[0]]
    header_map = {
        # Registration / Roll Number variations
        "registration number": "registration_number",
        "registration no": "registration_number",
        "reg no": "registration_number",
        "reg no ": "registration_number",
        "regno": "registration_number",
        "register number": "registration_number",
        "register no": "registration_number",
        "reg number": "registration_number",
        "student reg no": "registration_number",
        "student registration number": "registration_number",
        "student id": "registration_number",
        "roll number": "registration_number",
        "roll no": "registration_number",
        "rollno": "registration_number",
        "roll": "registration_number",
        "usn": "registration_number",

        # Student Name variations
        "name": "name",
        "full name": "name",
        "student name": "name",
        "candidate name": "name",
        "name of the student": "name",
        "name of candidate": "name",
        "student": "name",

        # Email variations
        "email": "email",
        "student email": "email",
        "email address": "email",
        "email id": "email",
        "emailid": "email",
        "mail id": "email",
        "mail": "email",
        "e mail": "email",
        "e mail id": "email",
        "student mail id": "email",

        # Phone / Mobile variations
        "phone": "phone",
        "mobile": "phone",
        "phone number": "phone",
        "mobile number": "phone",
        "contact": "phone",
        "contact no": "phone",
        "contact number": "phone",
        "mobile no": "phone",
        "phone no": "phone",

        # Department / Branch / Course variations
        "department": "department",
        "branch": "department",
        "dept": "department",
        "course": "department",
        "degree": "department",
        "discipline": "department",
        "stream": "department",
        "program": "department",
        "programme": "department",
        "specialization": "department",

        # CGPA / Academics
        "cgpa": "cgpa",
        "gpa": "cgpa",
        "overall cgpa": "cgpa",
        "current cgpa": "cgpa",
        "percentage": "academic_details",
        "marks": "academic_details",
        "score": "academic_details",
        "10th percentage": "academic_details",
        "12th percentage": "academic_details",
        "sslc": "academic_details",
        "hsc": "academic_details",
        "diploma percentage": "academic_details",
        "backlogs": "academic_details",
        "year": "academic_details",
        "skills": "skills",
        "technical skills": "skills",
        "resume link": "academic_details",
        "linkedin": "academic_details",
        "github": "academic_details",
        "github id": "academic_details",
        "placement status": "placement_status",
        "status": "placement_status",
        "placement": "placement_status",
        "company": "academic_details",
        "package": "offer_package_lpa",
        "salary": "offer_package_lpa",
        "ctc": "offer_package_lpa",
        "drive link": "academic_details",
        "gender": "academic_details",
        "student type": "academic_details",
    }

    lookup = {}
    for index, header in enumerate(headers):
        normalized = _normalize_excel_header(header)
        target = header_map.get(normalized)
        if target and target not in lookup:
            lookup[target] = index

    # Fallback to column index 0 (Reg No), 1 (Name), 2 (Dept/Email) if headers were not labeled standardly
    if "registration_number" not in lookup and len(headers) >= 1:
        lookup["registration_number"] = 0
    if "name" not in lookup and len(headers) >= 2:
        lookup["name"] = 1
    if "department" not in lookup and len(headers) >= 3:
        lookup["department"] = 2

    if "registration_number" not in lookup or "name" not in lookup:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Missing required columns: Registration Number and Student Name.")

    created = []
    updated = 0
    duplicates = 0
    invalid_rows = 0
    errors: list[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(_clean_excel_value(cell) for cell in row):
            continue

        registration_number = _clean_excel_value(row[lookup["registration_number"]]) if lookup.get("registration_number") is not None and lookup["registration_number"] < len(row) else ""
        name = _clean_excel_value(row[lookup["name"]]) if lookup.get("name") is not None and lookup["name"] < len(row) else ""
        email = _clean_excel_value(row[lookup["email"]]) if lookup.get("email") is not None and lookup["email"] < len(row) else ""
        department = _clean_excel_value(row[lookup["department"]]) if lookup.get("department") is not None and lookup["department"] < len(row) else "General"
        if not department:
            department = "General"

        if not registration_number:
            invalid_rows += 1
            errors.append(f"Row {row_number}: Missing registration number")
            continue

        normalized_registration = registration_number.strip()
        
        # Auto-generate email if not provided
        if not email or "@" not in email:
            email = f"{normalized_registration.lower().replace(' ', '.')}@student.rportal.local"
        
        email_lower = email.strip().lower()

        phone = _clean_excel_value(row[lookup["phone"]]) if "phone" in lookup and lookup["phone"] < len(row) else None
        phone = phone or None

        cgpa_raw = row[lookup["cgpa"]] if "cgpa" in lookup and lookup["cgpa"] < len(row) else None
        cgpa = None
        if cgpa_raw is not None and _clean_excel_value(cgpa_raw) != "":
            cgpa_value = _coerce_float(cgpa_raw)
            if cgpa_value is not None and 0 <= cgpa_value <= 10:
                cgpa = str(cgpa_value)

        placement_status = _parse_placement_status(row[lookup["placement_status"]] if "placement_status" in lookup and lookup["placement_status"] < len(row) else "")
        is_eligible = placement_status != PlacementStatus.NOT_ELIGIBLE

        academic_parts = []
        for field_name in ["academic_details", "cgpa", "phone"]:
            if field_name in lookup and lookup[field_name] < len(row):
                value = _clean_excel_value(row[lookup[field_name]])
                if value:
                    academic_parts.append(value)
        academic_details = "; ".join(academic_parts) if academic_parts else None

        skills_raw = row[lookup["skills"]] if "skills" in lookup and lookup["skills"] < len(row) else None
        skills = _clean_excel_value(skills_raw) if skills_raw is not None else None
        offer_package_lpa = _clean_excel_value(row[lookup["offer_package_lpa"]]) if "offer_package_lpa" in lookup and lookup["offer_package_lpa"] < len(row) else None

        existing = db.scalar(
            select(Student).where(
                (Student.registration_number == normalized_registration)
                | (Student.email == email_lower)
            )
        )

        if existing:
            if mode == "upsert":
                if name: existing.name = name
                if department: existing.department = department
                if phone: existing.phone = phone
                if cgpa: existing.cgpa = cgpa
                if skills: existing.skills = skills
                if academic_details: existing.academic_details = academic_details
                if offer_package_lpa: existing.offer_package_lpa = offer_package_lpa
                existing.placement_status = placement_status
                existing.is_eligible = is_eligible
                updated += 1
            else:
                duplicates += 1
                errors.append(f"Row {row_number}: Duplicate registration number ({normalized_registration}) skipped.")
            continue

        if not name:
            invalid_rows += 1
            errors.append(f"Row {row_number}: Missing student name")
            continue

        student = Student(
            registration_number=normalized_registration,
            name=name,
            email=email_lower,
            phone=phone,
            department=department,
            academic_details=academic_details,
            cgpa=cgpa,
            skills=skills,
            is_eligible=is_eligible,
            placement_status=placement_status,
            offer_package_lpa=offer_package_lpa or None,
            drive_links=[],
        )

        try:
            db.add(student)
            db.flush()
            created.append(student)
        except Exception:
            db.rollback()
            invalid_rows += 1
            errors.append(f"Row {row_number}: Database record validation failed")
            continue

    if not created and updated == 0:
        if duplicates > 0:
            return {
                "message": f"Processed {duplicates} record(s): All students already exist in the database (0 new records added).",
                "imported": 0,
                "updated": 0,
                "duplicates": duplicates,
                "invalid": invalid_rows,
                "errors": errors[:30],
                "total_processed": duplicates + invalid_rows,
            }
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "No valid student records could be imported.",
                "imported": 0,
                "updated": 0,
                "duplicates": duplicates,
                "invalid": invalid_rows,
                "errors": errors[:30],
            }
        )

    try:
        log_action(db, current, "IMPORT", "student", "batch", {"created": len(created), "updated": updated, "duplicates": duplicates})
        notify_admins_and_managers(db, "Students Imported", f"{len(created)} student(s) imported, {updated} updated by {current.full_name}.", "STUDENTS_IMPORTED", "student", "batch", exclude_user_id=current.id)
        db.commit()
    except Exception as exc:  # pragma: no cover
        db.rollback()
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "The import could not be saved. Please review the uploaded rows and try again.") from exc

    return {
        "message": f"Successfully imported {len(created)} student(s){f' and updated {updated} record(s)' if updated else ''}.",
        "imported": len(created),
        "updated": updated,
        "duplicates": duplicates,
        "invalid": invalid_rows,
        "errors": errors[:30],
        "total_processed": len(created) + updated + duplicates + invalid_rows,
    }


@app.get("/api/students", response_model=list[StudentOut])
def list_students(query: str | None = None, department: str | None = None, eligible: bool | None = None, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    stmt=select(Student).order_by(Student.name)
    if query: stmt=stmt.where((Student.name.ilike(f"%{query}%")) | (Student.email.ilike(f"%{query}%")) | (Student.registration_number.ilike(f"%{query}%")))
    if department: stmt=stmt.where(Student.department==department)
    if eligible is not None: stmt=stmt.where(Student.is_eligible==eligible)
    return db.scalars(stmt).all()


@app.get("/api/students/{student_id}", response_model=StudentOut)
def get_student(student_id:int,db:Session=Depends(get_db),current:User=Depends(get_current_user)):
    student=db.get(Student,student_id)
    if not student: raise HTTPException(404,"Student not found")
    return student


@app.post("/api/students",response_model=StudentOut,status_code=201)
def create_student(payload:StudentCreate,db:Session=Depends(get_db),current:User=Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    if db.scalar(select(Student).where((Student.email==payload.email)|(Student.registration_number==payload.registration_number))): raise HTTPException(409,"Student email or registration number already exists")
    student=Student(**payload.model_dump());db.add(student);db.flush()
    log_action(db,current,"CREATE","student",student.id,{"registration_number":student.registration_number,"name":student.name})
    notify_admins_and_managers(db, "New Student Added", f"Student {student.name} ({student.registration_number}) was added.", "STUDENT_CREATED", "student", student.id, exclude_user_id=current.id)
    audit_commit(db);db.refresh(student);return student


@app.put("/api/students/{student_id}",response_model=StudentOut)
def update_student(student_id:int,payload:StudentCreate,db:Session=Depends(get_db),current:User=Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    student=db.get(Student,student_id)
    if not student: raise HTTPException(404,"Student not found")
    for key,value in payload.model_dump().items(): setattr(student,key,value)
    log_action(db,current,"UPDATE","student",student.id,{"registration_number":student.registration_number,"name":student.name})
    notify_admins_and_managers(db, "Student Record Updated", f"Student {student.name} ({student.registration_number}) details were updated.", "STUDENT_UPDATED", "student", student.id, exclude_user_id=current.id)
    audit_commit(db);db.refresh(student);return student


@app.delete("/api/students/{student_id}",status_code=204)
def delete_student(student_id:int,db:Session=Depends(get_db),current:User=Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    student=db.get(Student,student_id)
    if not student: raise HTTPException(404,"Student not found")
    log_action(db,current,"DELETE","student",student.id,{"registration_number":student.registration_number,"name":student.name})
    notify_admins_and_managers(db, "Student Record Deleted", f"Student {student.name} ({student.registration_number}) was deleted.", "STUDENT_DELETED", "student", student.id, exclude_user_id=current.id)
    db.delete(student);audit_commit(db)


@app.get("/api/recruiters/overview", response_model=RecruitersOverviewOut)
def recruiters_overview(
    status: str | None = None,
    recruiter_status: str | None = None,
    company_id: int | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user)
):
    companies = db.scalars(select(Company).options(joinedload(Company.drives), joinedload(Company.recruiters)).order_by(Company.name)).unique().all()
    all_recruiters = db.scalars(select(RecruiterContact).options(joinedload(RecruiterContact.company)).order_by(RecruiterContact.name)).unique().all()
    all_drives = db.scalars(select(PlacementDrive).where(PlacementDrive.is_archived.is_(False))).all()
    
    total_recruiters = len(all_recruiters)
    active_recruiters_count = len([r for r in all_recruiters if r.status == "ACTIVE"])
    connected_companies_count = len(companies)
    total_drives_count = len(all_drives)
    active_drives_count = len([d for d in all_drives if d.status == DriveStatus.OPEN])
    completed_drives_count = len([d for d in all_drives if d.status == DriveStatus.CLOSED])

    cold_count = len([c for c in companies if c.recruiter_status == RecruiterStatus.COLD])
    warm_count = len([c for c in companies if c.recruiter_status == RecruiterStatus.WARM])
    hot_count = len([c for c in companies if c.recruiter_status == RecruiterStatus.HOT])
    drive_completed_count = len([c for c in companies if c.recruiter_status == RecruiterStatus.DRIVE_COMPLETED])

    company_cards = []
    for c in companies:
        c_drives = [d for d in all_drives if d.company_id == c.id]
        c_drive_ids = [d.id for d in c_drives]
        c_recruiters = [r for r in all_recruiters if r.company_id == c.id]
        
        apps_count = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids))) if c_drive_ids else 0
        shortlisted_count = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids), Application.status.in_([ApplicationStatus.SHORTLISTED, ApplicationStatus.INTERVIEW, ApplicationStatus.OFFERED]))) if c_drive_ids else 0
        selected_count = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids), Application.status == ApplicationStatus.OFFERED)) if c_drive_ids else 0
        
        primary_recruiter = c_recruiters[0] if c_recruiters else None
        latest_drive = sorted(c_drives, key=lambda d: d.created_at, reverse=True)[0] if c_drives else None

        company_cards.append(CompanyCardOut(
            id=c.id,
            name=c.name,
            website=c.website,
            industry=c.industry,
            contact_name=c.contact_name or (primary_recruiter.name if primary_recruiter else None),
            contact_email=c.contact_email or (primary_recruiter.email if primary_recruiter else None),
            contact_phone=c.contact_phone or (primary_recruiter.phone if primary_recruiter else None),
            contact_designation=c.contact_designation or (primary_recruiter.designation if primary_recruiter else "HR Manager"),
            logo_url=c.logo_url,
            notes=c.notes,
            last_contacted_at=c.last_contacted_at or (latest_drive.created_at if latest_drive else c.updated_at),
            recruiter_status=c.recruiter_status,
            created_at=c.created_at,
            updated_at=c.updated_at,
            recruiter_count=len(c_recruiters),
            primary_contact=primary_recruiter.name if primary_recruiter else c.contact_name,
            primary_email=primary_recruiter.email if primary_recruiter else c.contact_email,
            primary_phone=primary_recruiter.phone if primary_recruiter else c.contact_phone,
            primary_designation=primary_recruiter.designation if primary_recruiter else c.contact_designation,
            total_drives=len(c_drives),
            latest_drive_title=latest_drive.title if latest_drive else None,
            latest_drive_date=latest_drive.drive_date if latest_drive else None,
            latest_drive_status=latest_drive.status.value if latest_drive else None,
            applicants_count=apps_count or 0,
            shortlisted_count=shortlisted_count or 0,
            selected_count=selected_count or 0,
        ))

    recruiter_list = []
    for r in all_recruiters:
        r_comp = r.company
        c_drives = [d for d in all_drives if r_comp and d.company_id == r_comp.id]
        recruiter_list.append({
            "id": r.id,
            "company_id": r.company_id,
            "company_name": r_comp.name if r_comp else "—",
            "company_status": r_comp.recruiter_status.value if r_comp else "COLD",
            "company_industry": r_comp.industry if r_comp else "—",
            "company_logo": r_comp.logo_url if r_comp else None,
            "name": r.name,
            "designation": r.designation or "HR Manager",
            "email": r.email,
            "phone": r.phone or "—",
            "alternate_phone": r.alternate_phone or "—",
            "department": r.department or "Talent Acquisition",
            "linkedin_url": r.linkedin_url,
            "status": r.status,
            "last_contacted": r.last_contacted.isoformat() if r.last_contacted else None,
            "notes": r.notes,
            "total_drives": len(c_drives),
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })

    return RecruitersOverviewOut(
        summary={
            "total_recruiters": total_recruiters,
            "active_recruiters": active_recruiters_count,
            "connected_companies": connected_companies_count,
            "placement_drives": total_drives_count,
            "active_drives": active_drives_count,
            "completed_drives": completed_drives_count,
        },
        engagement_distribution={
            "cold": cold_count,
            "warm": warm_count,
            "hot": hot_count,
            "drive_completed": drive_completed_count,
        },
        companies=company_cards,
        recruiters=recruiter_list,
    )


@app.get("/api/recruiters/active", response_model=list[CompanyOut])
def active_recruiters(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    return db.scalars(select(Company).where(Company.recruiter_status.in_([RecruiterStatus.HOT, RecruiterStatus.WARM, RecruiterStatus.DRIVE_COMPLETED])).order_by(Company.recruiter_status, Company.name)).all()


@app.get("/api/recruiters/contacts", response_model=list[dict])
def list_recruiter_contacts(
    status: str | None = None,
    company_id: int | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user)
):
    query = select(RecruiterContact).options(joinedload(RecruiterContact.company)).order_by(RecruiterContact.name)
    if status and status != "ALL":
        query = query.where(RecruiterContact.status == status)
    if company_id:
        query = query.where(RecruiterContact.company_id == company_id)
    if search:
        query = query.join(Company).where(
            (RecruiterContact.name.ilike(f"%{search}%"))
            | (RecruiterContact.email.ilike(f"%{search}%"))
            | (RecruiterContact.phone.ilike(f"%{search}%"))
            | (RecruiterContact.designation.ilike(f"%{search}%"))
            | (Company.name.ilike(f"%{search}%"))
        )
    contacts = db.scalars(query).unique().all()
    results = []
    for r in contacts:
        drives_count = db.scalar(select(func.count()).select_from(PlacementDrive).where(PlacementDrive.company_id == r.company_id)) or 0
        results.append({
            "id": r.id,
            "company_id": r.company_id,
            "company_name": r.company.name if r.company else "—",
            "company_status": r.company.recruiter_status.value if r.company else "COLD",
            "company_industry": r.company.industry if r.company else "—",
            "company_logo": r.company.logo_url if r.company else None,
            "name": r.name,
            "designation": r.designation or "HR Manager",
            "email": r.email,
            "phone": r.phone or "—",
            "alternate_phone": r.alternate_phone or "—",
            "department": r.department or "Talent Acquisition",
            "linkedin_url": r.linkedin_url,
            "status": r.status,
            "last_contacted": r.last_contacted.isoformat() if r.last_contacted else None,
            "notes": r.notes,
            "total_drives": drives_count,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return results


@app.get("/api/recruiters", response_model=list[CompanyOut])
def list_recruiters(status: str | None = None, search: str | None = None, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    query = select(Company)
    if status == "ACTIVE":
        query = query.where(Company.recruiter_status.in_([RecruiterStatus.HOT, RecruiterStatus.WARM, RecruiterStatus.DRIVE_COMPLETED]))
    elif status and status != "ALL":
        try:
            enum_status = RecruiterStatus(status)
            query = query.where(Company.recruiter_status == enum_status)
        except ValueError:
            pass
    if search:
        query = query.where((Company.name.ilike(f"%{search}%")) | (Company.contact_name.ilike(f"%{search}%")) | (Company.contact_email.ilike(f"%{search}%")) | (Company.industry.ilike(f"%{search}%")))
    return db.scalars(query.order_by(Company.recruiter_status, Company.name)).all()


@app.get("/api/recruiters/sample-template")
def download_recruiter_template(format: str = Query("xlsx", pattern="^(xlsx|csv)$")):
    headers = ["Name", "Company", "Designation", "Email", "Phone", "Status", "Last Contacted", "Notes"]
    sample_rows = [
        ["Priya Sharma", "Tata Consultancy Services", "Campus HR Lead", "priya.sharma@tcs.com", "+91 9876543210", "ACTIVE", "2026-08-25", "Primary contact for Digital hiring."],
        ["Karthik Raja", "Zoho Corporation", "Lead Talent Partner", "campus@zohocorp.com", "+91 9876543211", "ACTIVE", "2026-08-28", "Coordinating SDE product drive."],
        ["Rohan Deshmukh", "Amazon Web Services", "University Relations Specialist", "aws-campus@amazon.com", "+91 9876543212", "ACTIVE", "2026-08-30", "Cloud support associate hiring."],
    ]
    if format == "csv":
        output = io.StringIO()
        import csv
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in sample_rows:
            writer.writerow(row)
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="recruiters_sample_template.csv"'}
        )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recruiters"
        ws.append(headers)
        for row in sample_rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="recruiters_sample_template.xlsx"'}
        )


@app.get("/api/recruiters/{recruiter_id}", response_model=CompanyDetailsOut)
@app.get("/api/companies/{recruiter_id}/details", response_model=CompanyDetailsOut)
def get_recruiter_details(recruiter_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    company = db.get(Company, recruiter_id)
    if not company:
        raise HTTPException(404, "Company not found")
    
    drives = db.scalars(select(PlacementDrive).where(PlacementDrive.company_id == recruiter_id).order_by(PlacementDrive.created_at.desc())).all()
    drive_ids = [d.id for d in drives]
    
    # 1. Recruiters
    recruiters = db.scalars(select(RecruiterContact).where(RecruiterContact.company_id == recruiter_id).order_by(RecruiterContact.name)).all()
    
    # 2. Applications for this company
    applications_raw = db.scalars(
        select(Application)
        .options(joinedload(Application.student), joinedload(Application.drive))
        .where(Application.drive_id.in_(drive_ids))
        .order_by(Application.created_at.desc())
    ).unique().all() if drive_ids else []
    
    total_applications = len(applications_raw)
    selected_students_count = len([a for a in applications_raw if a.status == ApplicationStatus.OFFERED])
    
    # 3. Placed students for this company
    placed_students_query = db.scalars(
        select(Student).where(Student.placed_company_id == recruiter_id)
    ).all()
    
    placed_students_list = []
    seen_placed_ids = set()
    for s in placed_students_query:
        seen_placed_ids.add(s.id)
        placed_students_list.append({
            "student_id": s.id,
            "name": s.name,
            "registration_number": s.registration_number,
            "department": s.department,
            "cgpa": s.cgpa or "—",
            "package_lpa": s.offer_package_lpa or "Best in Industry",
            "placed_date": s.updated_at.isoformat() if s.updated_at else None,
            "drive_title": "Campus Placement",
        })
        
    for a in applications_raw:
        if a.status == ApplicationStatus.OFFERED and a.student and a.student.id not in seen_placed_ids:
            seen_placed_ids.add(a.student.id)
            placed_students_list.append({
                "student_id": a.student.id,
                "name": a.student.name,
                "registration_number": a.student.registration_number,
                "department": a.student.department,
                "cgpa": a.student.cgpa or "—",
                "package_lpa": a.drive.package_lpa if a.drive else "Best in Industry",
                "placed_date": a.created_at.isoformat(),
                "drive_title": a.drive.title if a.drive else "Campus Drive",
            })
    
    # 4. Enriched Drives list
    drives_list = []
    jds_list = []
    for d in drives:
        d_apps = [a for a in applications_raw if a.drive_id == d.id]
        app_count = len(d_apps)
        shortlisted_count = len([a for a in d_apps if a.status == ApplicationStatus.SHORTLISTED])
        interview_count = len([a for a in d_apps if a.status == ApplicationStatus.INTERVIEW])
        offered_count = len([a for a in d_apps if a.status == ApplicationStatus.OFFERED])
        
        req_skills_list = [s.strip() for s in (d.required_skills or "").split(",") if s.strip()]
        pref_skills_list = [s.strip() for s in (d.preferred_skills or "").split(",") if s.strip()]
        dept_list = [dept.strip() for dept in (d.departments or "").split(",") if dept.strip()]
        
        drives_list.append({
            "id": d.id,
            "title": d.title,
            "job_role": d.job_role or d.title,
            "location": d.location,
            "package_lpa": d.package_lpa or "—",
            "eligibility": d.eligibility,
            "min_cgpa": d.min_cgpa or "6.0",
            "max_backlogs": d.max_backlogs or 0,
            "required_skills": d.required_skills,
            "preferred_skills": d.preferred_skills,
            "departments": d.departments,
            "status": d.status.value,
            "work_mode": d.work_mode or "On-site",
            "drive_date": d.drive_date.isoformat() if d.drive_date else None,
            "deadline": d.deadline.isoformat() if d.deadline else None,
            "applications_count": app_count,
            "shortlisted_count": shortlisted_count,
            "interview_count": interview_count,
            "offered_count": offered_count,
            "placed_count": offered_count,
            "created_at": d.created_at.isoformat() if d.created_at else None,
        })
        
        jds_list.append({
            "id": d.id,
            "drive_id": d.id,
            "drive_title": d.title,
            "job_title": d.title,
            "job_role": d.job_role or d.title,
            "description": d.description or d.eligibility,
            "required_skills": req_skills_list,
            "preferred_skills": pref_skills_list,
            "eligible_departments": dept_list,
            "min_cgpa": d.min_cgpa or "6.0",
            "max_backlogs": d.max_backlogs or 0,
            "experience_requirement": d.experience_requirement or "Fresher / Final Year",
            "certifications": d.certifications or "Not Required",
            "package_lpa": d.package_lpa or "As per company norms",
            "location": d.location,
            "jd_document_path": d.jd_document_path,
            "created_at": d.created_at.isoformat() if d.created_at else None,
        })
    
    # 5. Eligible Students calculation
    eligible_students_list = []
    if drives:
        all_students = db.scalars(select(Student).where(Student.is_eligible.is_(True)).order_by(Student.name)).all()
        for s in all_students[:30]:
            eligible_students_list.append({
                "id": s.id,
                "name": s.name,
                "registration_number": s.registration_number,
                "department": s.department,
                "cgpa": s.cgpa or "—",
                "skills": s.skills or "—",
                "placement_status": s.placement_status.value,
            })
    
    # 6. Applications detailed list
    apps_list = []
    for a in applications_raw:
        apps_list.append({
            "id": a.id,
            "student_id": a.student_id,
            "student_name": a.student_name,
            "student_email": a.student_email,
            "department": a.student.department if a.student else "—",
            "cgpa": a.student.cgpa if a.student else "—",
            "drive_id": a.drive_id,
            "drive_title": a.drive.title if a.drive else "—",
            "status": a.status.value,
            "applied_at": a.created_at.isoformat() if a.created_at else None,
            "resume_path": a.resume_path,
        })
        
    # 7. Activity History (audit logs)
    audit_logs = db.scalars(
        select(AuditLog)
        .where(
            (AuditLog.entity_type == "company") & (AuditLog.entity_id == str(recruiter_id))
            | (AuditLog.entity_type == "placement_drive") & (AuditLog.entity_id.in_([str(did) for did in drive_ids]))
        )
        .order_by(AuditLog.created_at.desc())
        .limit(25)
    ).all()
    
    activity_history_list = [{
        "id": log.id,
        "action": log.action,
        "entity_type": log.entity_type,
        "entity_id": log.entity_id,
        "details": log.details,
        "created_at": log.created_at.isoformat(),
    } for log in audit_logs]
    
    active_drives_count = len([d for d in drives if d.status == DriveStatus.OPEN and not d.is_archived])
    last_drive = max([d.created_at for d in drives], default=None) if drives else company.updated_at
    
    return CompanyDetailsOut(
        id=company.id,
        name=company.name,
        website=company.website,
        industry=company.industry,
        location=company.location,
        address=company.address,
        description=company.description,
        contact_name=company.contact_name,
        contact_email=company.contact_email,
        contact_phone=company.contact_phone,
        contact_designation=company.contact_designation,
        logo_url=company.logo_url,
        notes=company.notes,
        last_contacted_at=company.last_contacted_at,
        recruiter_status=company.recruiter_status,
        created_at=company.created_at,
        updated_at=company.updated_at,
        total_recruiters=len(recruiters),
        total_drives=len(drives),
        active_drives=active_drives_count,
        total_applications=total_applications,
        selected_students=selected_students_count,
        placed_students_count=len(placed_students_list),
        last_engagement=last_drive,
        recruiters=[RecruiterContactOut.model_validate(r) for r in recruiters],
        drives=drives_list,
        job_descriptions=jds_list,
        eligible_students=eligible_students_list,
        applications=apps_list,
        placed_students=placed_students_list,
        activity_history=activity_history_list,
    )


@app.patch("/api/companies/{company_id}/status", response_model=CompanyOut)
def update_company_status(company_id: int, payload: CompanyStatusUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))):
    company = db.get(Company, company_id)
    if not company:
        raise HTTPException(404, "Company not found")
    old_status = company.recruiter_status.value
    company.recruiter_status = payload.status
    company.last_contacted_at = datetime.now(timezone.utc)
    
    log_action(db, current, "COMPANY_STATUS_CHANGED", "company", company.id, {"company": company.name, "from": old_status, "to": payload.status.value})
    notify_admins_and_managers(
        db,
        "Company Engagement Status Updated",
        f"{company.name} status updated to {payload.status.value} by {current.full_name}.",
        "RECRUITER_STATUS_CHANGED",
        "company",
        company.id,
        exclude_user_id=current.id
    )
    audit_commit(db)
    db.refresh(company)
    return company


@app.post("/api/recruiters/contacts", response_model=dict, status_code=201)
def create_recruiter_contact(
    payload: RecruiterContactCreate,
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    company = db.get(Company, payload.company_id)
    if not company:
        raise HTTPException(404, "Target company does not exist")
    
    existing = db.scalar(select(RecruiterContact).where(RecruiterContact.company_id == payload.company_id, RecruiterContact.email == payload.email))
    if existing:
        raise HTTPException(409, f"Recruiter contact with email '{payload.email}' already exists for {company.name}")
    
    recruiter = RecruiterContact(**payload.model_dump())
    if not recruiter.last_contacted:
        recruiter.last_contacted = datetime.now(timezone.utc)
    db.add(recruiter)
    db.flush()

    if not company.contact_name or not company.contact_email:
        company.contact_name = recruiter.name
        company.contact_email = recruiter.email
        company.contact_phone = recruiter.phone
        company.contact_designation = recruiter.designation

    log_action(db, current, "CREATE", "recruiter_contact", recruiter.id, {"name": recruiter.name, "company": company.name})
    notify_admins_and_managers(
        db,
        "New Recruiter Added",
        f"Recruiter {recruiter.name} ({recruiter.designation or 'HR'}) from {company.name} added.",
        "RECRUITER_CREATED",
        "recruiter_contact",
        recruiter.id,
        exclude_user_id=current.id
    )
    audit_commit(db)
    db.refresh(recruiter)
    return {
        "id": recruiter.id,
        "name": recruiter.name,
        "company_id": company.id,
        "company_name": company.name,
        "email": recruiter.email,
        "designation": recruiter.designation,
        "status": recruiter.status,
        "message": f"Recruiter {recruiter.name} added successfully."
    }


@app.put("/api/recruiters/contacts/{contact_id}", response_model=dict)
@app.patch("/api/recruiters/contacts/{contact_id}", response_model=dict)
def update_recruiter_contact(
    contact_id: int,
    payload: RecruiterContactUpdate,
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    recruiter = db.get(RecruiterContact, contact_id)
    if not recruiter:
        raise HTTPException(404, "Recruiter contact not found")
    
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(recruiter, key, value)
    
    log_action(db, current, "UPDATE", "recruiter_contact", recruiter.id, {"name": recruiter.name, "email": recruiter.email})
    audit_commit(db)
    db.refresh(recruiter)
    return {
        "id": recruiter.id,
        "name": recruiter.name,
        "email": recruiter.email,
        "status": recruiter.status,
        "message": "Recruiter details updated successfully."
    }


@app.delete("/api/recruiters/contacts/{contact_id}", status_code=204)
def delete_recruiter_contact(
    contact_id: int,
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    recruiter = db.get(RecruiterContact, contact_id)
    if not recruiter:
        raise HTTPException(404, "Recruiter contact not found")
    
    company_name = recruiter.company.name if recruiter.company else "Company"
    log_action(db, current, "DELETE", "recruiter_contact", recruiter.id, {"name": recruiter.name, "company": company_name})
    notify_admins_and_managers(
        db,
        "Recruiter Contact Deleted",
        f"Recruiter {recruiter.name} ({company_name}) was deleted by {current.full_name}.",
        "RECRUITER_DELETED",
        "recruiter_contact",
        recruiter.id,
        exclude_user_id=current.id
    )
    db.delete(recruiter)
    audit_commit(db)


@app.post("/api/recruiters", response_model=CompanyOut, status_code=201)
def create_recruiter(payload: CompanyBase, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    if db.scalar(select(Company).where(Company.name == payload.name)):
        raise HTTPException(409, "A recruiter/company with this name already exists")
    company = Company(**payload.model_dump())
    db.add(company)
    db.flush()

    if company.contact_name and company.contact_email:
        recruiter = RecruiterContact(
            company_id=company.id,
            name=company.contact_name,
            designation=company.contact_designation or "HR Manager",
            email=company.contact_email,
            phone=company.contact_phone or "+91 98765 43210",
            department="Talent Acquisition",
            status="ACTIVE",
            last_contacted=datetime.now(timezone.utc),
            notes=f"Primary contact for {company.name}"
        )
        db.add(recruiter)

    log_action(db, current, "CREATE", "recruiter", company.id, {"name": company.name, "status": company.recruiter_status.value})
    notify_admins_and_managers(db, "New Recruiter Added", f"Recruiter/Company {company.name} was registered.", "RECRUITER_CREATED", "company", company.id, exclude_user_id=current.id)
    audit_commit(db)
    db.refresh(company)
    return company


@app.put("/api/recruiters/{recruiter_id}", response_model=CompanyOut)
def update_recruiter(recruiter_id: int, payload: CompanyBase, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    company = db.get(Company, recruiter_id)
    if not company:
        raise HTTPException(404, "Recruiter not found")
    old_status = company.recruiter_status.value
    for key, value in payload.model_dump().items():
        setattr(company, key, value)
    log_action(db, current, "UPDATE", "recruiter", company.id, {"name": company.name, "old_status": old_status, "new_status": company.recruiter_status.value})
    if old_status != company.recruiter_status.value:
        notify_admins_and_managers(db, "Recruiter Status Updated", f"{company.name} engagement status changed from {old_status} to {company.recruiter_status.value}.", "RECRUITER_STATUS_CHANGED", "company", company.id, exclude_user_id=current.id)
    audit_commit(db)
    db.refresh(company)
    return company


@app.delete("/api/recruiters/{recruiter_id}", status_code=204)
def delete_recruiter(recruiter_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    company = db.get(Company, recruiter_id)
    if not company:
        raise HTTPException(404, "Recruiter not found")
    log_action(db, current, "DELETE", "recruiter", company.id, {"name": company.name})
    notify_admins_and_managers(db, "Recruiter Deleted", f"Recruiter/Company {company.name} was deleted.", "RECRUITER_DELETED", "company", company.id, exclude_user_id=current.id)
    db.delete(company)
    audit_commit(db)


@app.post("/api/recruiters/import")
def import_recruiters(
    file: UploadFile = File(...),
    mode: str = Query("skip", pattern="^(skip|upsert)$"),
    db: Session = Depends(get_db),
    current: User = Depends(require_roles(Role.ADMIN, Role.MANAGER))
):
    if not file.filename:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No file selected.")

    ext = Path(file.filename).suffix.lower()
    if ext not in {".xlsx", ".xls", ".csv"}:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx, .xls, and .csv files are supported.")

    try:
        contents = file.file.read()
        rows = _parse_excel_rows(contents, file.filename)
    except Exception as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Invalid import file: {exc}") from exc

    if not rows or len(rows) < 2:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Import file has no data rows.")

    headers = [_clean_excel_value(cell) for cell in rows[0]]
    header_map = {
        "name": "name", "recruiter name": "name", "contact name": "name", "full name": "name",
        "company": "company", "company name": "company", "organization": "company",
        "designation": "designation", "role": "designation", "title": "designation", "job title": "designation",
        "email": "email", "recruiter email": "email", "contact email": "email", "email address": "email",
        "phone": "phone", "phone number": "phone", "mobile": "phone", "contact phone": "phone",
        "status": "status", "recruiter status": "status",
        "last contacted": "last_contacted", "last contact": "last_contacted",
        "notes": "notes", "comments": "notes", "remarks": "notes"
    }
    lookup = {}
    for idx, h in enumerate(headers):
        clean_h = h.strip().lower()
        if clean_h in header_map:
            lookup[header_map[clean_h]] = idx

    if "name" not in lookup or "company" not in lookup or "email" not in lookup:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Missing required columns: Name, Company, and Email.")

    created = 0
    updated = 0
    duplicates = 0
    invalid_rows = 0
    errors = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        
        name = _clean_excel_value(row[lookup["name"]]) if lookup["name"] < len(row) else ""
        company_name = _clean_excel_value(row[lookup["company"]]) if lookup["company"] < len(row) else ""
        email = _clean_excel_value(row[lookup["email"]]).strip().lower() if lookup["email"] < len(row) else ""
        
        if not name or not company_name or not email or "@" not in email:
            invalid_rows += 1
            errors.append(f"Row {row_number}: Missing name, company or invalid email.")
            continue

        company = db.scalar(select(Company).where(Company.name.ilike(company_name.strip())))
        if not company:
            company = Company(
                name=company_name.strip(),
                industry="IT & Technology",
                recruiter_status=RecruiterStatus.COLD
            )
            db.add(company)
            db.flush()

        designation = _clean_excel_value(row[lookup["designation"]]) if "designation" in lookup and lookup["designation"] < len(row) else "HR Manager"
        phone = _clean_excel_value(row[lookup["phone"]]) if "phone" in lookup and lookup["phone"] < len(row) else None
        rec_status = _clean_excel_value(row[lookup["status"]]).upper() if "status" in lookup and lookup["status"] < len(row) else "ACTIVE"
        if rec_status not in ["ACTIVE", "INACTIVE"]:
            rec_status = "ACTIVE"
        notes = _clean_excel_value(row[lookup["notes"]]) if "notes" in lookup and lookup["notes"] < len(row) else None

        existing_contact = db.scalar(select(RecruiterContact).where(RecruiterContact.email == email))
        if existing_contact:
            if mode == "upsert":
                existing_contact.name = name
                existing_contact.company_id = company.id
                if designation: existing_contact.designation = designation
                if phone: existing_contact.phone = phone
                existing_contact.status = rec_status
                if notes: existing_contact.notes = notes
                updated += 1
            else:
                duplicates += 1
                errors.append(f"Row {row_number}: Recruiter {email} already exists (skipped).")
            continue

        new_contact = RecruiterContact(
            company_id=company.id,
            name=name,
            designation=designation or "HR Manager",
            email=email,
            phone=phone or None,
            status=rec_status,
            notes=notes or None,
            last_contacted=datetime.now(timezone.utc)
        )
        db.add(new_contact)
        created += 1

    try:
        log_action(db, current, "IMPORT", "recruiter", "batch", {"created": created, "updated": updated, "duplicates": duplicates})
        notify_admins_and_managers(db, "Recruiters Imported", f"{created} recruiter(s) imported, {updated} updated by {current.full_name}.", "RECRUITERS_IMPORTED", "recruiter", "batch", exclude_user_id=current.id)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Import could not be saved. Please review data and try again.") from exc

    return {
        "message": f"Successfully imported {created} recruiter(s){f' and updated {updated} record(s)' if updated else ''}.",
        "imported": created,
        "updated": updated,
        "duplicates": duplicates,
        "invalid": invalid_rows,
        "errors": errors[:25],
        "total_processed": created + updated + duplicates + invalid_rows,
    }

def _team_member_payload(member: PlacementTeamMember):
    assigned_drives = []
    for assignment in sorted(member.drive_assignments, key=lambda item: item.created_at, reverse=True):
        if assignment.drive is not None:
            assigned_drives.append({
                "id": assignment.drive.id,
                "title": assignment.drive.title,
                "company": assignment.drive.company.name if assignment.drive.company else "—",
                "drive_date": assignment.drive.drive_date.isoformat() if assignment.drive.drive_date else None,
                "status": assignment.drive.status.value if assignment.drive.status else None,
                "responsibility": assignment.responsibility,
            })
    return {
        "id": member.id,
        "user_id": member.user_id,
        "role": member.role,
        "responsibility": member.responsibility,
        "department": member.department,
        "phone": member.phone,
        "assignment": member.assignment,
        "is_active": member.is_active,
        "is_team_lead": member.is_team_lead,
        "invitation_status": member.invitation_status,
        "joined_date": member.joined_date,
        "created_at": member.created_at,
        "updated_at": member.updated_at,
        "user": member.user,
        "assigned_drives": assigned_drives,
    }

@app.post("/api/placement-team/import")
def import_placement_team(file: UploadFile = File(...), db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    if not file.filename or Path(file.filename).suffix.lower() not in {".xlsx", ".xls"}:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx and .xls files are supported.")
    try:
        rows = _parse_excel_rows(file.file.read(), file.filename)
    except Exception as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Invalid Excel file: {exc}") from exc

    headers = [_normalize_excel_header(cell) for cell in rows[0]]
    aliases = {
        "email": {"email", "user email", "email address"},
        "role": {"role", "team role"},
        "responsibility": {"responsibility", "responsibility area"},
        "department": {"department", "team department"},
        "phone": {"phone", "phone number", "mobile"},
        "assignment": {"assignment", "assignment name"},
        "company": {"company", "company name"},
        "drive": {"drive", "drive title", "placement drive"},
        "team_lead": {"team lead", "is team lead", "lead"},
        "is_active": {"active", "is active", "status"},
    }
    lookup = {target: next((index for index, header in enumerate(headers) if header in names), None) for target, names in aliases.items()}
    if lookup["email"] is None or lookup["responsibility"] is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Missing required columns: Email and Responsibility.")

    imported = 0
    duplicates = 0
    invalid = 0
    errors = []

    def cell(row, field):
        index = lookup[field]
        return _clean_excel_value(row[index]) if index is not None and index < len(row) else ""

    def truthy(value, default=True):
        normalized = value.lower()
        if not normalized:
            return default
        return normalized not in {"false", "0", "no", "inactive", "in-active"}

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(_clean_excel_value(value) for value in row):
            continue
        email = cell(row, "email").lower()
        responsibility = cell(row, "responsibility")
        user = db.scalar(select(User).where(User.email == email))
        if not user or not responsibility:
            invalid += 1
            errors.append(f"Row {row_number} - Existing user email and responsibility are required")
            continue
        if db.scalar(select(PlacementTeamMember).where(PlacementTeamMember.user_id == user.id)):
            duplicates += 1
            continue

        member = PlacementTeamMember(
            user_id=user.id,
            role=cell(row, "role") or "Placement Officer",
            responsibility=responsibility,
            department=cell(row, "department") or None,
            phone=cell(row, "phone") or None,
            assignment=cell(row, "assignment") or None,
            is_active=truthy(cell(row, "is_active")),
            is_team_lead=truthy(cell(row, "team_lead"), False),
            invited_by_id=current.id,
            invitation_status="ACTIVE" if truthy(cell(row, "is_active")) else "INACTIVE",
        )
        db.add(member)
        db.flush()

        company_name = cell(row, "company")
        drive_title = cell(row, "drive")
        if company_name or drive_title:
            drives = db.scalars(select(PlacementDrive).options(joinedload(PlacementDrive.company)).order_by(PlacementDrive.created_at.desc())).unique().all()
            matching_drive = next((drive for drive in drives if (not company_name or (drive.company and drive.company.name.lower() == company_name.lower())) and (not drive_title or drive.title.lower() == drive_title.lower())), None)
            if matching_drive:
                db.add(PlacementTeamDriveAssignment(team_member_id=member.id, drive_id=matching_drive.id, responsibility=cell(row, "assignment") or responsibility))
            else:
                errors.append(f"Row {row_number} - No matching existing company/drive assignment found")
        imported += 1

    if not imported:
        db.rollback()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, {"message": "Import failed", "imported": 0, "duplicates": duplicates, "invalid": invalid, "errors": errors[:20]})
    log_action(db, current, "IMPORT", "placement_team_member", "batch", {"count": imported})
    notify_admins_and_managers(db, "Placement Team Imported", f"{imported} team members imported by {current.full_name}.", "TEAM_IMPORTED", "placement_team_member", "batch", exclude_user_id=current.id)
    db.commit()
    return {"message": "Import completed", "imported": imported, "duplicates": duplicates, "invalid": invalid, "errors": errors[:20]}


@app.get("/api/placement-team", response_model=list[TeamMemberOut])
def list_placement_team(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    members = db.scalars(
        select(PlacementTeamMember)
        .options(
            joinedload(PlacementTeamMember.user),
            joinedload(PlacementTeamMember.drive_assignments).joinedload(PlacementTeamDriveAssignment.drive).joinedload(PlacementDrive.company),
        )
        .order_by(PlacementTeamMember.created_at.desc())
    ).unique().all()
    return [_team_member_payload(member) for member in members]


@app.get("/api/team", response_model=list[TeamMemberOut])
def list_team(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    return list_placement_team(db=db, current=current)


@app.get("/api/placement-team/{member_id}", response_model=TeamMemberOut)
def get_placement_team_member(member_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    member = db.scalar(select(PlacementTeamMember).options(joinedload(PlacementTeamMember.user), joinedload(PlacementTeamMember.drive_assignments).joinedload(PlacementTeamDriveAssignment.drive).joinedload(PlacementDrive.company)).where(PlacementTeamMember.id == member_id))
    if not member:
        raise HTTPException(404, "Team member not found")
    return _team_member_payload(member)


@app.post("/api/placement-team", response_model=TeamMemberOut, status_code=201)
def add_team_member(payload: TeamMemberCreate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    target_user = db.get(User, payload.user_id)
    if not target_user:
        raise HTTPException(404, "User not found")
    if db.scalar(select(PlacementTeamMember).where(PlacementTeamMember.user_id == payload.user_id)):
        raise HTTPException(409, "User is already in the placement team")
    member = PlacementTeamMember(
        user_id=payload.user_id,
        role=payload.role,
        responsibility=payload.responsibility,
        department=payload.department,
        phone=payload.phone,
        assignment=payload.assignment,
        is_active=payload.is_active,
        is_team_lead=payload.is_team_lead,
        invited_by_id=current.id,
        invitation_status="ACTIVE",
    )
    db.add(member)
    db.flush()
    log_action(db, current, "CREATE", "placement_team_member", member.id, {"name": target_user.full_name, "role": payload.role})
    notify_admins_and_managers(db, "New Placement Team Member", f"{target_user.full_name} added as {payload.role}.", "TEAM_MEMBER_ADDED", "placement_team_member", member.id, exclude_user_id=current.id)
    audit_commit(db)
    return _team_member_payload(db.scalar(select(PlacementTeamMember).options(joinedload(PlacementTeamMember.user), joinedload(PlacementTeamMember.drive_assignments).joinedload(PlacementTeamDriveAssignment.drive).joinedload(PlacementDrive.company)).where(PlacementTeamMember.id == member.id)))


@app.post("/api/team", response_model=TeamMemberOut, status_code=201)
def add_team_member_legacy(payload: TeamMemberCreate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    return add_team_member(payload=payload, db=db, current=current)


@app.patch("/api/placement-team/{member_id}", response_model=TeamMemberOut)
def update_team_member(member_id: int, payload: TeamMemberUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    member = db.get(PlacementTeamMember, member_id)
    if not member:
        raise HTTPException(404, "Team member not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(member, key, value)
    if payload.is_active is not None:
        member.invitation_status = "ACTIVE" if payload.is_active else "INACTIVE"
    name = member.user.full_name if member.user else "team_member"
    log_action(db, current, "UPDATE", "placement_team_member", member.id, {"name": name, "status": member.is_active})
    notify_admins_and_managers(db, "Team Member Updated", f"{name}'s profile was updated.", "TEAM_MEMBER_UPDATED", "placement_team_member", member.id, exclude_user_id=current.id)
    audit_commit(db)
    return _team_member_payload(db.scalar(select(PlacementTeamMember).options(joinedload(PlacementTeamMember.user), joinedload(PlacementTeamMember.drive_assignments).joinedload(PlacementTeamDriveAssignment.drive).joinedload(PlacementDrive.company)).where(PlacementTeamMember.id == member.id)))


@app.patch("/api/team/{member_id}", response_model=TeamMemberOut)
def update_team_member_legacy(member_id: int, payload: TeamMemberUpdate, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    return update_team_member(member_id=member_id, payload=payload, db=db, current=current)


@app.get("/api/placement-team/{member_id}/drives", response_model=list[dict])
def get_member_drive_assignments(member_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    member = db.get(PlacementTeamMember, member_id)
    if not member:
        raise HTTPException(404, "Team member not found")
    return [{
        "id": assignment.id,
        "drive_id": assignment.drive_id,
        "title": assignment.drive.title if assignment.drive else None,
        "company": assignment.drive.company.name if assignment.drive and assignment.drive.company else None,
        "responsibility": assignment.responsibility,
        "status": assignment.drive.status.value if assignment.drive and assignment.drive.status else None,
    } for assignment in member.drive_assignments]


@app.post("/api/placement-team/{member_id}/drives", response_model=dict, status_code=201)
def assign_drive_to_member(member_id: int, payload: dict, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    member = db.get(PlacementTeamMember, member_id)
    if not member:
        raise HTTPException(404, "Team member not found")
    drive_id = payload.get("drive_id")
    drive = db.get(PlacementDrive, drive_id)
    if not drive:
        raise HTTPException(404, "Placement drive not found")
    if db.scalar(select(PlacementTeamDriveAssignment).where(PlacementTeamDriveAssignment.team_member_id == member_id, PlacementTeamDriveAssignment.drive_id == drive_id)):
        raise HTTPException(409, "This drive is already assigned to the team member")
    assignment = PlacementTeamDriveAssignment(team_member_id=member_id, drive_id=drive_id, responsibility=payload.get("responsibility") or member.responsibility)
    db.add(assignment)
    db.flush()
    log_action(db, current, "ASSIGN_DRIVE", "placement_team_member", member.id, {"drive_id": drive_id, "drive_title": drive.title})
    notify_admins_and_managers(db, "Drive Assigned to Member", f"{drive.title} assigned to {member.user.full_name if member.user else 'team member'}.", "DRIVE_ASSIGNED", "placement_team_member", member.id, exclude_user_id=current.id)
    audit_commit(db)
    return {"message": "Drive assigned successfully", "assignment_id": assignment.id, "team_member_id": member_id, "drive_id": drive_id}


@app.delete("/api/placement-team/{member_id}/drives/{drive_id}", status_code=204)
def remove_drive_from_member(member_id: int, drive_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN, Role.LEAD))):
    assignment = db.scalar(select(PlacementTeamDriveAssignment).where(PlacementTeamDriveAssignment.team_member_id == member_id, PlacementTeamDriveAssignment.drive_id == drive_id))
    if not assignment:
        raise HTTPException(404, "Drive assignment not found")
    log_action(db, current, "REMOVE_DRIVE_ASSIGNMENT", "placement_team_member", member_id, {"drive_id": drive_id})
    db.delete(assignment)
    audit_commit(db)


@app.get("/api/notifications",response_model=list[NotificationOut])
def notifications(db:Session=Depends(get_db),current:User=Depends(get_current_user)):
    return db.scalars(select(Notification).where(Notification.recipient_id==current.id).order_by(Notification.created_at.desc())).all()


@app.get("/api/notifications/unread-count")
def notification_count(db:Session=Depends(get_db),current:User=Depends(get_current_user)):
    return {"unread_count":db.scalar(select(func.count()).select_from(Notification).where(Notification.recipient_id==current.id,Notification.is_read.is_(False))) or 0}


@app.patch("/api/notifications/{notification_id}/read",response_model=NotificationOut)
def mark_notification_read(notification_id:int,db:Session=Depends(get_db),current:User=Depends(get_current_user)):
    note=db.get(Notification,notification_id)
    if not note or note.recipient_id!=current.id:raise HTTPException(404,"Notification not found")
    from datetime import datetime, timezone
    note.is_read=True;note.read_at=datetime.now(timezone.utc);audit_commit(db);return note

@app.patch("/api/notifications/read-all", response_model=dict)
def mark_all_notifications_read(db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)
    unread_notes = db.scalars(select(Notification).where(Notification.recipient_id == current.id, Notification.is_read.is_(False))).all()
    for note in unread_notes:
        note.is_read = True
        note.read_at = now
    audit_commit(db)
    return {"message": "All notifications marked as read", "count": len(unread_notes)}


@app.get("/api/reports", response_model=ReportsOut)
def reports(
    query: str | None = None,
    company: str | None = None,
    department: str | None = None,
    student_status: str | None = None,
    company_status: str | None = None,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user)
):
    total = db.scalar(select(func.count()).select_from(Student)) or 0
    eligible = db.scalar(select(func.count()).select_from(Student).where(Student.is_eligible.is_(True))) or 0
    placed = db.scalar(select(func.count()).select_from(Student).where(Student.placement_status == PlacementStatus.PLACED)) or 0
    unplaced = db.scalar(select(func.count()).select_from(Student).where(Student.placement_status == PlacementStatus.SEEKING)) or 0
    
    departments = [{"department": name, "placed": count} for name, count in db.execute(
        select(Student.department, func.count()).where(Student.placement_status == PlacementStatus.PLACED).group_by(Student.department)
    ).all()]
    
    funnel = [{"status": status.value, "count": count} for status, count in db.execute(
        select(Application.status, func.count()).group_by(Application.status)
    ).all()]
    
    cold = db.scalar(select(func.count()).select_from(Company).where(Company.recruiter_status == RecruiterStatus.COLD)) or 0
    warm = db.scalar(select(func.count()).select_from(Company).where(Company.recruiter_status == RecruiterStatus.WARM)) or 0
    hot = db.scalar(select(func.count()).select_from(Company).where(Company.recruiter_status == RecruiterStatus.HOT)) or 0
    drive_completed = db.scalar(select(func.count()).select_from(Company).where(Company.recruiter_status == RecruiterStatus.DRIVE_COMPLETED)) or 0
    
    total_companies = db.scalar(select(func.count()).select_from(Company)) or 0
    total_applications = db.scalar(select(func.count()).select_from(Application)) or 0
    offers = db.scalar(select(func.count()).select_from(Application).where(Application.status == ApplicationStatus.OFFERED)) or 0
    active_drives = db.scalar(select(func.count()).select_from(PlacementDrive).where(PlacementDrive.status == DriveStatus.OPEN, PlacementDrive.is_archived.is_(False))) or 0

    # Recruiter metrics
    companies_list = db.scalars(select(Company).order_by(Company.name)).all()
    recruiter_metrics = []
    for comp in companies_list:
        c_drives = db.scalars(select(PlacementDrive).where(PlacementDrive.company_id == comp.id)).all()
        c_drive_ids = [d.id for d in c_drives]
        c_apps = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids))) if c_drive_ids else 0
        c_selected = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids), Application.status == ApplicationStatus.OFFERED)) if c_drive_ids else 0
        c_rejected = db.scalar(select(func.count()).select_from(Application).where(Application.drive_id.in_(c_drive_ids), Application.status == ApplicationStatus.REJECTED)) if c_drive_ids else 0
        c_active_drives = len([d for d in c_drives if d.status == DriveStatus.OPEN and not d.is_archived])
        last_drive = max([d.created_at for d in c_drives], default=None) if c_drives else comp.updated_at
        recruiter_metrics.append(RecruiterMetricsOut(
            company_id=comp.id,
            company_name=comp.name,
            contact_name=comp.contact_name,
            contact_email=comp.contact_email,
            status=comp.recruiter_status,
            total_drives=len(c_drives),
            active_drives=c_active_drives,
            total_applications=c_apps or 0,
            selected_count=c_selected or 0,
            rejected_count=c_rejected or 0,
            last_engagement=last_drive
        ))

    # Build detailed filterable report records
    student_query = select(Student).options(joinedload(Student.placed_company)).order_by(Student.name)
    if department and department != "ALL":
        student_query = student_query.where(Student.department == department)
    if student_status and student_status != "ALL":
        if student_status.upper() == "PLACED":
            student_query = student_query.where(Student.placement_status == PlacementStatus.PLACED)
        elif student_status.upper() in ["UNPLACED", "SEEKING"]:
            student_query = student_query.where(Student.placement_status == PlacementStatus.SEEKING)
        elif student_status.upper() == "NOT_ELIGIBLE":
            student_query = student_query.where(Student.placement_status == PlacementStatus.NOT_ELIGIBLE)
    if query:
        student_query = student_query.where((Student.name.ilike(f"%{query}%")) | (Student.registration_number.ilike(f"%{query}%")) | (Student.email.ilike(f"%{query}%")))
    
    students_data = db.scalars(student_query).unique().all()
    
    records = []
    for st in students_data:
        comp_name = st.placed_company.name if st.placed_company else None
        comp_status = st.placed_company.recruiter_status.value if st.placed_company else None
        
        if company and company != "ALL":
            if not comp_name or comp_name.lower() != company.lower():
                continue
        if company_status and company_status != "ALL":
            if not comp_status or comp_status.upper() != company_status.upper():
                continue

        apps_count = db.scalar(select(func.count()).select_from(Application).where(Application.student_id == st.id)) or 0
        
        records.append({
            "id": st.id,
            "registration_number": st.registration_number,
            "student_name": st.name,
            "email": st.email,
            "department": st.department,
            "cgpa": st.cgpa or "—",
            "student_status": st.placement_status.value,
            "company_name": comp_name or "—",
            "company_status": comp_status or "—",
            "package_lpa": st.offer_package_lpa or "—",
            "applied_drives_count": apps_count,
            "created_at": st.created_at.isoformat() if st.created_at else None,
        })

    return ReportsOut(
        total_students=total,
        eligible_students=eligible,
        placed_students=placed,
        unplaced_students=unplaced,
        placement_percentage=round(placed * 100 / total, 2) if total else 0,
        total_companies=total_companies,
        cold_recruiters=cold,
        warm_recruiters=warm,
        hot_recruiters=hot,
        drive_completed_recruiters=drive_completed,
        applications=total_applications,
        offers=offers,
        active_drives=active_drives,
        department_placements=departments,
        application_funnel=funnel,
        recruiter_metrics=recruiter_metrics,
        records=records
    )


# Serve frontend static files
from fastapi.staticfiles import StaticFiles
@app.delete("/api/placement-team/{member_id}", status_code=204)
def delete_team_member(member_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    member = db.get(PlacementTeamMember, member_id)
    if not member:
        raise HTTPException(404, "Team member not found")
    db.delete(member)
    log_action(db, current, "DELETE", "placement_team_member", member_id, {})
    audit_commit(db)
    return None


@app.delete("/api/applications/{application_id}", status_code=204)
def delete_application(application_id: int, db: Session = Depends(get_db), current: User = Depends(require_roles(Role.ADMIN))):
    application = db.get(Application, application_id)
    if not application:
        raise HTTPException(404, "Application not found")
    db.delete(application)
    log_action(db, current, "DELETE", "application", application_id, {})
    audit_commit(db)
    return None


from fastapi.responses import FileResponse

frontend_dir = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "frontend", "dist"
)

if os.path.exists(frontend_dir):
    assets_dir = os.path.join(frontend_dir, "assets")
    if os.path.exists(assets_dir):
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")

    @app.get("/{catchall:path}")
    def serve_frontend(catchall: str):
        if catchall.startswith("api/"):
            raise HTTPException(status_code=404, detail="Not Found")
        
        file_path = os.path.join(frontend_dir, catchall)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
            
        return FileResponse(os.path.join(frontend_dir, "index.html"))

